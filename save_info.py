# import xlwt
# from openpyxl import load_workbook
import openpyxl
from openpyxl import load_workbook

zm = ['a', 'b', 'c', 'd', 'e', 'f', 'g', 'h', 'i', 'j', 'k', 'l', 'm', 'n', 'o', 'p', 'q', 'r', 's', 't', 'u', 'v', 'w',
      'x', 'y', 'z']


def save_Info(sheetHeader, data_info, title):
    wb = openpyxl.Workbook()
    sheet = wb.active
    for i in range(len(sheetHeader)):
        sheet[zm[i] + '1'] = sheetHeader[i]
    for i in range(len(data_info)):
        num = 1
        for k in data_info[i]:
            sheet.cell(row=i + 2, column=num).value = data_info[i][k]
            num = num + 1
    wb.save(title)
    print('数据写入成功！')

def save_Info1(sheetHeader, data_info, title):
        wb = openpyxl.Workbook()
        sheet = wb.active
        for i in range(len(sheetHeader)):
            sheet[zm[i] + '1'] = sheetHeader[i]


        for i in range(len(data_info)):
            print(i)
            num = 1
            for k in data_info[i]:
                print(k)
                sheet.cell(row=i + 2, column=num).value = k
                num = num + 1
        wb.save(title)
        print('数据写入成功！')
    # # 加载excel，注意路径要与脚本一致
    # wb = load_workbook('demo.xlsx')
    # # 激活excel表
    # sheet = wb.active
    #
    # # 向excel中写入表头
    # sheet['a1'] = '姓名'
    # sheet['b1'] = '性别'
    # sheet['c1'] = '年龄'
    #
    # # 向excel中写入对应的value
    # sheet.cell(row=2, column=1).value = '张三'
    # sheet.cell(row=2, column=2).value = '男'
    # sheet.cell(row=2, column=3).value = 20
    #
    # wb.save('demo.xlsx')
    # print('数据写入成功！')

    # Workbook = xlwt.Workbook(encoding='utf-8')
    # Sheet_Name = Workbook.add_sheet('运维家')
    # Headers = ['姓名']
    # for index, Header in enumerate(Headers):
    #     Sheet_Name.write(0, index, Header)
    #
    # Names = ['张三', '李四', '王五']
    #
    # for index, Name in enumerate(Names):
    #     Sheet_Name.write(index + 1, 0, Name)
    #
    # print('=====')
    # Workbook.save('ceshi2.xls')
